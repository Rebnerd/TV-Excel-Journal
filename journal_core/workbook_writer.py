"""Excel workbook section writers for the daily report template."""

from __future__ import annotations

from collections import defaultdict
from copy import copy as _copy
from datetime import date, datetime, time
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter

from .config import (
    CHART_DATA_COLS,
    CHART_DATA_START_ROW,
    CURVE_COLS,
    CURVE_START_ROW,
    DURATION_BOUNDARY_PAD_SECONDS,
    DURATION_COLS,
    DURATION_START_ROW,
    DEFAULT_SUMMARY_CELLS,
    TRADE_COLS,
    TZ_CME,
    TZ_LOCAL,
)
from .time_utils import (
    floor_datetime_to_minute,
    format_duration_label,
    parse_amp_datetime,
    parse_duration_to_seconds,
)


def require_sheet(wb, sheet_name: str):
    """Return a workbook sheet or raise a clear error if the template is missing it."""
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Could not find sheet: {sheet_name}")
    return wb[sheet_name]

def clear_sheet_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(r,c).value = None

def _trade_entry_time_to_local_minute(entry_time_value) -> Optional[datetime]:
    """Map a trade entry timestamp onto the Chart Data timeline in Asia/Shanghai time.

    Trades are stored as local (Asia/Shanghai) timestamps in this report. The chart
    timeline should also be local time, so we only need to normalize the timezone and
    truncate to minute precision.
    """
    if entry_time_value is None:
        return None
    if isinstance(entry_time_value, datetime):
        dt_local = entry_time_value.replace(tzinfo=TZ_LOCAL) if entry_time_value.tzinfo is None else entry_time_value.astimezone(TZ_LOCAL)
    else:
        try:
            dt_local = parse_amp_datetime(str(entry_time_value))
        except Exception:
            return None
    return floor_datetime_to_minute(dt_local).replace(tzinfo=None)

def classify_trade_entry_series(trade: Dict[str, Any]) -> Optional[str]:
    direction = str(trade.get("direction") or "").strip().lower()
    pnl = float(trade.get("pnl_usd") or 0.0)
    if pnl == 0:
        return None
    if direction == "long":
        return "long_win" if pnl > 0 else "long_loss"
    if direction == "short":
        return "short_win" if pnl > 0 else "short_loss"
    return None

def build_chart_data_rows(close_df: pd.DataFrame, trades: List[Dict]) -> List[Dict[str, Any]]:
    if close_df is None or close_df.empty:
        return []

    df = close_df.copy()
    if "time_ct" not in df.columns or "close" not in df.columns:
        raise ValueError("close_df must contain columns: time_ct, close")

    def _to_excel_local_minute(v):
        if pd.isna(v):
            return None
        if isinstance(v, pd.Timestamp):
            py = v.to_pydatetime()
        else:
            py = v
        if not isinstance(py, datetime):
            py = pd.to_datetime(py).to_pydatetime()
        if py.tzinfo is not None:
            py = py.astimezone(TZ_LOCAL)
        else:
            # fetch_full_session_1m_close returns time_ct; if tz is missing, treat it as CT first
            py = py.replace(tzinfo=TZ_CME).astimezone(TZ_LOCAL)
        return floor_datetime_to_minute(py).replace(tzinfo=None)

    df["__time_excel"] = df["time_ct"].apply(_to_excel_local_minute)
    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    df = df.dropna(subset=["__time_excel", "close"]).copy()
    df = df.sort_values("__time_excel").drop_duplicates(subset=["__time_excel"], keep="last")

    trade_points_by_minute: Dict[datetime, List[Tuple[str, float, datetime, Any]]] = defaultdict(list)
    for tr in trades:
        series_key = classify_trade_entry_series(tr)
        if series_key is None:
            continue
        entry_minute = _trade_entry_time_to_local_minute(tr.get("entry_time"))
        if entry_minute is None:
            continue
        entry_price = tr.get("entry_price")
        try:
            entry_price = float(entry_price)
        except Exception:
            continue
        raw_entry_time = tr.get("entry_time")
        sort_entry = raw_entry_time if isinstance(raw_entry_time, datetime) else datetime.min
        trade_points_by_minute[entry_minute].append((series_key, entry_price, sort_entry, tr.get("id")))

    rows: List[Dict[str, Any]] = []
    for _, close_row in df.iterrows():
        minute_key = close_row["__time_excel"]
        close_val = float(close_row["close"])
        minute_trades = trade_points_by_minute.get(minute_key, [])
        if not minute_trades:
            rows.append({
                "time": minute_key,
                "close": close_val,
                "long_win": None,
                "long_loss": None,
                "short_win": None,
                "short_loss": None,
            })
            continue

        minute_trades = sorted(minute_trades, key=lambda x: (x[2], x[3]))
        for series_key, entry_price, _, _ in minute_trades:
            row = {
                "time": minute_key,
                "close": close_val,
                "long_win": None,
                "long_loss": None,
                "short_win": None,
                "short_loss": None,
            }
            row[series_key] = entry_price
            rows.append(row)

    return rows

def get_chart_data_rth_split_rows(
    chart_rows: List[Dict[str, Any]],
    cme_trade_date: date,
) -> Tuple[int, int]:
    """Return (eth_end_row, rth_start_row) for the close chart split at 08:30 CT.

    The Chart Data sheet itself stays unchanged:
      - A = local (Asia/Shanghai) time
      - B = close
      - C..F = trade-entry dot series

    We only split the *chart series ranges*:
      - ETH Close uses A/B rows before the 08:30 CT boundary
      - RTH Close uses A/B rows from the 08:30 CT boundary onward

    Returned rows are Excel row numbers and are always clamped to valid ranges.
    """
    if not chart_rows:
        return CHART_DATA_START_ROW, CHART_DATA_START_ROW

    local_open_naive = (
        datetime.combine(cme_trade_date, time(8, 30), tzinfo=TZ_CME)
        .astimezone(TZ_LOCAL)
        .replace(tzinfo=None)
    )

    first_rth_index: Optional[int] = None
    for i, row in enumerate(chart_rows):
        tval = row.get("time")
        if isinstance(tval, datetime) and tval >= local_open_naive:
            first_rth_index = i
            break

    last_row = CHART_DATA_START_ROW + len(chart_rows) - 1

    if first_rth_index is None:
        # Fallback: no RTH boundary found in the written rows; keep both valid.
        return last_row, last_row

    rth_start_row = CHART_DATA_START_ROW + first_rth_index
    eth_end_row = rth_start_row - 1

    if eth_end_row < CHART_DATA_START_ROW:
        eth_end_row = CHART_DATA_START_ROW
    if rth_start_row > last_row:
        rth_start_row = last_row

    return eth_end_row, rth_start_row

def _normalize_market_time_to_ct(value) -> Optional[datetime]:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        dt_obj = value.to_pydatetime()
    else:
        dt_obj = value
    if not isinstance(dt_obj, datetime):
        try:
            dt_obj = pd.to_datetime(dt_obj).to_pydatetime()
        except Exception:
            return None
    if dt_obj.tzinfo is None:
        return dt_obj.replace(tzinfo=TZ_CME)
    return dt_obj.astimezone(TZ_CME)

def _ohlc_from_close_rows(rows: pd.DataFrame) -> Optional[Dict[str, float]]:
    if rows is None or rows.empty:
        return None
    close_values = pd.to_numeric(rows["close"], errors="coerce").dropna()
    if close_values.empty:
        return None
    return {
        "open": float(close_values.iloc[0]),
        "high": float(close_values.max()),
        "low": float(close_values.min()),
        "close": float(close_values.iloc[-1]),
    }

def _range_metrics(ohlc: Optional[Dict[str, float]], prior_close: Optional[float]) -> Dict[str, Optional[float]]:
    if not ohlc:
        return {"range": None, "range_pct": None, "true_range": None}
    range_value = float(ohlc["high"]) - float(ohlc["low"])
    close_value = float(ohlc["close"])
    true_range = None
    if prior_close is not None:
        true_range = max(
            range_value,
            abs(float(ohlc["high"]) - float(prior_close)),
            abs(float(ohlc["low"]) - float(prior_close)),
        )
    return {
        "range": range_value,
        "range_pct": (range_value / close_value if close_value != 0 else None),
        "true_range": true_range,
    }

def build_market_summary(close_df: pd.DataFrame, cme_trade_date: date, prior_close: Optional[float]) -> Dict[str, Any]:
    """Build Summary market stats from one symbol's fetched 1-minute close data."""
    if close_df is None or close_df.empty or "time_ct" not in close_df.columns or "close" not in close_df.columns:
        return {}

    df = close_df.copy()
    df["__time_ct"] = df["time_ct"].apply(_normalize_market_time_to_ct)
    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    df = df.dropna(subset=["__time_ct", "close"]).sort_values("__time_ct").copy()
    if df.empty:
        return {}

    rth_start = datetime.combine(cme_trade_date, time(8, 30), tzinfo=TZ_CME)
    full_ohlc = _ohlc_from_close_rows(df)
    eth_ohlc = _ohlc_from_close_rows(df[df["__time_ct"] < rth_start])
    rth_ohlc = _ohlc_from_close_rows(df[df["__time_ct"] >= rth_start])
    full_range = _range_metrics(full_ohlc, prior_close)
    eth_range = _range_metrics(eth_ohlc, prior_close)
    rth_range = _range_metrics(rth_ohlc, prior_close)

    out: Dict[str, Any] = {}
    if full_ohlc:
        out.update(full_ohlc)
        out["direction"] = "Up" if full_ohlc["close"] >= full_ohlc["open"] else "Down"
    out.update({
        "range": full_range["range"],
        "range_pct": full_range["range_pct"],
        "true_range": full_range["true_range"],
        "eth_range": eth_range["range"],
        "eth_range_pct": eth_range["range_pct"],
        "eth_true_range": eth_range["true_range"],
        "rth_range": rth_range["range"],
        "rth_range_pct": rth_range["range_pct"],
        "rth_true_range": rth_range["true_range"],
    })
    return out

def write_chart_data(ws_chart_data, chart_rows: List[Dict[str, Any]]) -> int:
    max_existing = ws_chart_data.max_row
    if max_existing >= CHART_DATA_START_ROW:
        clear_sheet_range(ws_chart_data, CHART_DATA_START_ROW, max_existing, 1, 6)

    headers = {
        "time": "time",
        "close": "close",
        "long_win": "Long Win",
        "long_loss": "Long Loss",
        "short_win": "Short Win",
        "short_loss": "Short Loss",
    }
    for key, col_idx in CHART_DATA_COLS.items():
        ws_chart_data.cell(1, col_idx).value = headers[key]

    for i, row in enumerate(chart_rows):
        r = CHART_DATA_START_ROW + i
        for key, col_idx in CHART_DATA_COLS.items():
            value = row.get(key)
            cell = ws_chart_data.cell(r, col_idx)
            cell.value = value
            if key == "time" and value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif key != "time" and value is not None:
                cell.number_format = "0.00"

    end_row = CHART_DATA_START_ROW + max(0, len(chart_rows) - 1)
    if len(chart_rows) == 0:
        end_row = CHART_DATA_START_ROW
    return end_row

def copy_row_style(ws, src_row: int, dst_row: int, min_col: int, max_col: int):
    """Copy the *template* row styling cell-by-cell.
    
    Important: we must clone the underlying style object per destination cell.
    Otherwise, later border edits (used for partial-exit outlining) would leak
    to other rows because they'd share the same style instance.
    """
    for c in range(min_col, max_col+1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        dst._style = _copy(src._style)
        dst.number_format = src.number_format

def write_trades(ws_trades, trades: List[Dict]):
    """
    Write trades into the Trades sheet (locked column order).

    Visual grouping (per your request):
      - If consecutive rows share the same (entry_time, entry_price) AND the group has >1 rows,
        insert ONE blank row **before** the group and ONE blank row **after** the group
        (except at the very top / very bottom), so the transition into/out of a partial-exit
        group is visually separated.

    Notes:
      - This is purely a visual separator (no changes to trade matching / P&L logic).
      - We insert rows AFTER writing, and insert from bottom-to-top to avoid index shifting issues.
    """
    # Clear old data (rows 2..max_row) AND reset styles so previous runs don't linger.
    max_existing = ws_trades.max_row
    prototype_row = 2  # style prototype row in the template (first data row)

    if max_existing >= 2:
        for r in range(2, max_existing + 1):
            for c in range(1, len(TRADE_COLS) + 1):
                ws_trades.cell(r, c).value = None

        # Reset styles for existing rows
        if prototype_row <= ws_trades.max_row:
            for r in range(2, max_existing + 1):
                copy_row_style(ws_trades, prototype_row, r, 1, len(TRADE_COLS))

    # Write rows with prototype style
    for i, tr in enumerate(trades):
        row_idx = 2 + i

        # Copy style from prototype row if it exists
        if prototype_row <= ws_trades.max_row:
            copy_row_style(ws_trades, prototype_row, row_idx, 1, len(TRADE_COLS))

        # Write values in locked order
        for col_i, key in enumerate(TRADE_COLS, start=1):
            ws_trades.cell(row_idx, col_i).value = tr.get(key)

    # Insert blank rows around partial-exit groups (same entry_time + same entry_price, group size > 1)
    # Implementation detail: we insert *rows* (not borders). To insert a blank row *before* a group
    # starting at Excel row R, we insert a row *after* Excel row (R-1).
    insert_after_rows_set = set()
    if trades:
        # 1) Build consecutive groups
        groups: List[Tuple[int, int]] = []  # (start_i, end_i), inclusive indices in `trades`
        start_i = 0
        current_key = (trades[0].get("entry_time"), trades[0].get("entry_price"))
        for i in range(1, len(trades)):
            key = (trades[i].get("entry_time"), trades[i].get("entry_price"))
            if key != current_key:
                groups.append((start_i, i - 1))
                start_i = i
                current_key = key
        groups.append((start_i, len(trades) - 1))

        # 2) For each multi-row group, add a blank row before and after (if not at edges)
        for gs, ge in groups:
            if ge - gs + 1 <= 1:
                continue

            # Blank row BEFORE the group (if it is not the very first trade row)
            if gs > 0:
                # Insert after the row of the trade immediately above group start
                # Trade index (gs-1) is at Excel row: 2 + (gs-1)
                insert_after_rows_set.add(2 + (gs - 1))

            # Blank row AFTER the group (if it is not the final trade row)
            if ge < len(trades) - 1:
                # Trade index ge is at Excel row: 2 + ge
                insert_after_rows_set.add(2 + ge)

    insert_after_rows: List[int] = sorted(insert_after_rows_set)

    # Perform insertions from bottom to top so row indices remain valid
    for end_row in sorted(insert_after_rows, reverse=True):
        insert_row = end_row + 1
        ws_trades.insert_rows(insert_row, 1)

        # Style the inserted blank row to match the template's data row style
        if prototype_row <= ws_trades.max_row:
            copy_row_style(ws_trades, prototype_row, insert_row, 1, len(TRADE_COLS))

        # Ensure the inserted row is truly blank
        for c in range(1, len(TRADE_COLS) + 1):
            ws_trades.cell(insert_row, c).value = None

    # Update autofilter range to include all rows (including inserted blanks)
    last_row = ws_trades.max_row
    ws_trades.auto_filter.ref = f"A1:{get_column_letter(len(TRADE_COLS))}{last_row}"

    # Add an outline rectangle around each partial-exit group (same entry_time + entry_price, size>1)
    # so it is visually obvious at a glance (purely formatting; no logic changes).
    outline_side = Side(style="medium", color="000000")
    entry_time_col = TRADE_COLS.index("entry_time") + 1
    entry_price_col = TRADE_COLS.index("entry_price") + 1
    col_start, col_end = 1, len(TRADE_COLS)

    def _apply_outline(r0: int, r1: int) -> None:
        # Top and bottom borders across the full row width
        for c in range(col_start, col_end + 1):
            top_cell = ws_trades.cell(r0, c)
            bot_cell = ws_trades.cell(r1, c)
            top_cell.border = top_cell.border.copy(top=outline_side)
            bot_cell.border = bot_cell.border.copy(bottom=outline_side)
        # Left and right borders along the group height
        for r in range(r0, r1 + 1):
            left_cell = ws_trades.cell(r, col_start)
            right_cell = ws_trades.cell(r, col_end)
            left_cell.border = left_cell.border.copy(left=outline_side)
            right_cell.border = right_cell.border.copy(right=outline_side)

    current_key = None
    group_start = None
    group_size = 0
    for r in range(2, last_row + 1):
        et = ws_trades.cell(r, entry_time_col).value
        ep = ws_trades.cell(r, entry_price_col).value
        # Treat fully blank spacer rows as hard separators
        if et is None and ep is None:
            if group_start is not None and group_size > 1:
                _apply_outline(group_start, r - 1)
            current_key = None
            group_start = None
            group_size = 0
            continue

        key = (et, ep)
        if group_start is None:
            group_start = r
            current_key = key
            group_size = 1
        elif key == current_key:
            group_size += 1
        else:
            if group_start is not None and group_size > 1:
                _apply_outline(group_start, r - 1)
            group_start = r
            current_key = key
            group_size = 1

    if group_start is not None and group_size > 1:
        _apply_outline(group_start, last_row)


    # ---------------------------------------------------------
    # ETH ↔ RTH divider (blue line) — no blank row insertion
    # ---------------------------------------------------------
    blue_side = Side(style="medium", color="0000FF")
    entry_time_col = TRADE_COLS.index("entry_time") + 1
    col_start, col_end = 1, len(TRADE_COLS)

    def _is_rth_entry_time(et_val):
        """Treat sheet entry_time as Asia/Shanghai naive datetime; convert to America/Chicago and check RTH."""
        if et_val is None:
            return None
        # openpyxl writes naive datetimes; interpret as TZ_LOCAL
        if isinstance(et_val, datetime):
            dt_local = et_val.replace(tzinfo=TZ_LOCAL) if et_val.tzinfo is None else et_val.astimezone(TZ_LOCAL)
        else:
            try:
                dt_local = parse_amp_datetime(str(et_val))
            except Exception:
                return None
        dt_ct = dt_local.astimezone(TZ_CME)
        tt = dt_ct.time()
        return (tt >= time(8, 30) and tt < time(15, 0))

    # Collect real trade rows (skip spacer rows)
    trade_rows = []  # list of (row_idx, is_rth_bool)
    for r in range(2, last_row + 1):
        et = ws_trades.cell(r, entry_time_col).value
        if et is None:
            continue
        flag = _is_rth_entry_time(et)
        if flag is None:
            continue
        trade_rows.append((r, flag))

    # Apply a blue top border on the first row after a session transition
    for i in range(1, len(trade_rows)):
        prev_is_rth = trade_rows[i - 1][1]
        curr_row, curr_is_rth = trade_rows[i]
        if prev_is_rth != curr_is_rth:
            for c in range(col_start, col_end + 1):
                cell = ws_trades.cell(curr_row, c)
                cell.border = cell.border.copy(top=blue_side)

def write_summary(
    ws_summary,
    local_date: date,
    cme_trade_date: date,
    market_summary_by_symbol: Optional[Dict[str, Dict[str, Any]]],
    metrics: Dict[str, float],
    summary_cells: Optional[Dict[str, Any]] = None,
):
    cells = summary_cells or DEFAULT_SUMMARY_CELLS
    ws_summary[cells["local_date"]].value = local_date.strftime("%Y-%m-%d")
    ws_summary[cells["cme_trade_date"]].value = cme_trade_date.strftime("%Y-%m-%d")
    if "commission_per_side" in cells:
        ws_summary[cells["commission_per_side"]].value = None

    for symbol, market_cells in cells.get("market_blocks", {}).items():
        symbol_summary = (market_summary_by_symbol or {}).get(symbol, {})
        for key, addr in market_cells.items():
            cell = ws_summary[addr]
            cell.value = symbol_summary.get(key)
            if key.endswith("_range_pct"):
                cell.number_format = "0.00%"
            elif key != "direction":
                cell.number_format = "0.00"

    for k, addr in cells.items():
        if k in (
            "local_date",
            "cme_trade_date",
            "commission_per_side",
            "market_blocks",
        ):
            continue
        if k in metrics:
            ws_summary[addr].value = metrics[k]

def write_curve(ws_curve, trades: List[Dict], cme_trade_date: date):
    usd_number_format = "$#,##0.00;-$#,##0.00"

    # Build cumulative PnL series ordered by exit_time (local)
    trades_sorted = sorted(trades, key=lambda d: (d["exit_time"], d["id"]))
    cum = 0.0
    rows = []
    for tr in trades_sorted:
        cum += float(tr.get("pnl_usd") or 0.0)
        t = tr["exit_time"]
        # time label as HH:MM:SS (force Asia/Shanghai display)
        if getattr(t, 'tzinfo', None) is None:
            t_local = t.replace(tzinfo=TZ_LOCAL)
        else:
            t_local = t.astimezone(TZ_LOCAL)
        time_label = t_local.strftime("%H:%M:%S")
        cum_pos = cum if cum >= 0 else None
        cum_neg = cum if cum < 0 else None
        rows.append((time_label, float(round(cum,2)), cum_pos, cum_neg))

    # Clear existing curve data area
    max_existing = ws_curve.max_row
    if max_existing >= CURVE_START_ROW:
        clear_sheet_range(ws_curve, CURVE_START_ROW, max_existing, 1, 4)

    # Write rows
    for i, (time_label, cum_pnl, cum_pos, cum_neg) in enumerate(rows):
        r = CURVE_START_ROW + i
        ws_curve.cell(r, CURVE_COLS["time_label"]).value = time_label
        for key, value in (
            ("cum_pnl", cum_pnl),
            ("cum_pos", cum_pos),
            ("cum_neg", cum_neg),
        ):
            cell = ws_curve.cell(r, CURVE_COLS[key])
            cell.value = value
            cell.number_format = usd_number_format

    # Update the session title text (row 2 in template)
    ws_curve["A2"].value = f"Session: CME trade date {cme_trade_date.strftime('%Y-%m-%d')} (17:00–16:00 CT). X=Local time, Y=Cumulative P&L (USD)."

    # Return the last data row for later chart-XML patching (post-save).
    end_row = CURVE_START_ROW + max(0, len(rows) - 1)
    if len(rows) == 0:
        end_row = CURVE_START_ROW
    return end_row

def write_duration_distribution(ws_duration, trades: List[Dict]) -> int:
    """Write exact-duration helper data with right-boundary protective spacing.

    Output columns / signs:
      A: Duration
      B: Winning Amount ($) -> positive (upside)
      C: Losing Amount ($)  -> negative (downside)
      D: Winning MFE ($)    -> positive (upside)
      E: Winning MAE ($)    -> negative (downside)
      F: Losing MFE ($)     -> positive (upside)
      G: Losing MAE ($)     -> negative (downside)

    The helper table is zero-padded second-by-second all the way to the max duration,
    plus an extra right-side boundary pad so the last real bar is not glued to the chart edge.
    """
    aggregates: Dict[int, Dict[str, float]] = defaultdict(lambda: {
        "winning_amount_usd": 0.0,
        "losing_amount_usd": 0.0,
        "winning_mfe_usd": 0.0,
        "winning_mae_usd": 0.0,
        "losing_mfe_usd": 0.0,
        "losing_mae_usd": 0.0,
    })

    max_sec = 0
    for tr in trades:
        sec = parse_duration_to_seconds(tr.get("duration"))
        if sec is None:
            continue

        max_sec = max(max_sec, sec)
        pnl = float(tr.get("pnl_usd") or 0.0)
        mfe = float(tr.get("mfe_usd") or 0.0)
        mae = float(tr.get("mae_usd") or 0.0)

        if pnl > 0:
            aggregates[sec]["winning_amount_usd"] += abs(pnl)
            aggregates[sec]["winning_mfe_usd"] += abs(mfe)
            aggregates[sec]["winning_mae_usd"] += -abs(mae)
        elif pnl < 0:
            aggregates[sec]["losing_amount_usd"] += -abs(pnl)
            aggregates[sec]["losing_mfe_usd"] += abs(mfe)
            aggregates[sec]["losing_mae_usd"] += -abs(mae)

    max_existing = ws_duration.max_row
    if max_existing >= DURATION_START_ROW:
        clear_sheet_range(ws_duration, DURATION_START_ROW, max_existing, 1, 7)

    headers = {
        "duration_label": "Duration",
        "winning_amount_usd": "Winning Amount ($)",
        "losing_amount_usd": "Losing Amount ($)",
        "winning_mfe_usd": "Winning MFE ($)",
        "winning_mae_usd": "Winning MAE ($)",
        "losing_mfe_usd": "Losing MFE ($)",
        "losing_mae_usd": "Losing MAE ($)",
    }
    for key, col_idx in DURATION_COLS.items():
        ws_duration.cell(1, col_idx).value = headers[key]

    padded_max_sec = max_sec + int(DURATION_BOUNDARY_PAD_SECONDS)

    for sec in range(0, padded_max_sec + 1):
        r = DURATION_START_ROW + sec
        bucket = aggregates.get(sec, {})
        ws_duration.cell(r, DURATION_COLS["duration_label"]).value = format_duration_label(sec)
        for key in (
            "winning_amount_usd",
            "losing_amount_usd",
            "winning_mfe_usd",
            "winning_mae_usd",
            "losing_mfe_usd",
            "losing_mae_usd",
        ):
            val = float(bucket.get(key, 0.0))
            # Keep empty buckets visually blank in DurationDist, like the earlier version.
            # Only write a number when this duration actually has non-zero aggregated data.
            ws_duration.cell(r, DURATION_COLS[key]).value = round(val, 10) if val != 0 else None

    end_row = DURATION_START_ROW + padded_max_sec
    return end_row
